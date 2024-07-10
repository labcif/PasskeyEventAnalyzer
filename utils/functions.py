import csv
import pandas as pd
import openpyxl
import openpyxl.utils
import openpyxl.utils.dataframe
import os
from scripts.ilapfuncs import *


def write_csv(file_name, data):  # data is a matrix
    with open(file_name, "w", newline='', encoding='utf-8') as f:
        wr = csv.writer(f)
        wr.writerows(data)


def read_csv(file_name):
    with open(file_name, "r", encoding='utf-8') as f:
        rd = csv.reader(f)
        return list(rd)

def write_excel(file_name, sheet, data, is_rewrite=False):

    df = pd.DataFrame(data)
    
    if is_rewrite:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook[workbook.sheetnames[0]]) # remove default sheet
        workbook.create_sheet(sheet)
        
        for i in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=False):
            workbook[sheet].append(i)

        workbook.save(file_name)
    else:
        try:
            workbook = openpyxl.load_workbook(file_name)
            
            if sheet in workbook.sheetnames:
                workbook.remove(workbook[sheet])
                workbook.create_sheet(sheet)
            else:
                workbook.create_sheet(sheet)
            
            for i in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=False):
                workbook[sheet].append(i)

            workbook.save(file_name)

        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook[workbook.sheetnames[0]]) # remove default sheet
            workbook.create_sheet(sheet)
            
            for i in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=False):
                workbook[sheet].append(i)

            workbook.save(file_name)

